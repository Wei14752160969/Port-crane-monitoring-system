import matplotlib.pyplot as plt
import sys
import openpyxl
import random
import matplotlib
matplotlib.use("Qt5Agg")
from PyQt5.QtWidgets import QVBoxLayout, QSizePolicy, QWidget, QApplication
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure


class MyMplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=5, height=4, dpi=100):

        # 配置中文显示
        plt.rcParams['font.family'] = ['SimHei']  # 用来正常显示中文标签
        plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

        self.fig = Figure(figsize=(width, height), dpi=dpi)  # 新建一个figure
        # self.axes = self.fig.add_subplot(111)  # 建立一个子图，如果要建立复合图，可以在这里修改
        # self.axes.cla()  # 每次绘图的时候不保留上一次绘图的结果

        FigureCanvas.__init__(self, self.fig)
        self.setParent(parent)

        '''定义FigureCanvas的尺寸策略，这部分的意思是设置FigureCanvas，使之尽可能的向外填充空间。'''
        FigureCanvas.setSizePolicy(self,
                                   QSizePolicy.Expanding,
                                   QSizePolicy.Expanding)
        FigureCanvas.updateGeometry(self)

    # 绘制图形
    def angle_2_plot(self):
        # 创建workbook
        wb5 = openpyxl.Workbook()
        sheet = wb5.active
        sheet.append(['时间', '数据'])
        # 生成数据
        x1 = []
        y1 = []
        for i in range(0, 60):
            nowTime1 = i
            randomNum1 = random.random() - 0.5
            x1.append(nowTime1)
            y1.append(randomNum1)
            sheet.append([str(nowTime1), str(randomNum1)[0:8]])
        wb5.save('倾角5.xlsx')
        ax = self.figure.add_subplot(221)
        ax.plot(x1[0:60], y1[0:60])
        ax.set_ylabel('角度')


        # 创建workbook
        wb6 = openpyxl.Workbook()
        sheet = wb6.active
        sheet.append(['时间', '数据'])
        # 生成数据
        x2 = []
        y2 = []
        for i in range(0, 60):
            nowTime2 = i
            randomNum2 = random.random() - 0.5
            x2.append(nowTime2)
            y2.append(randomNum2)
            sheet.append([str(nowTime2)[0:8], str(randomNum2)[0:8]])
        wb6.save('倾角6.xlsx')
        bx = self.figure.add_subplot(222)
        bx.plot(x2[0:60], y2[0:60])
        bx.set_ylabel('角度')


        # 创建workbook
        wb7 = openpyxl.Workbook()
        sheet = wb7.active
        sheet.append(['时间', '数据'])
        # 生成数据
        x3 = []
        y3 = []
        for i in range(0, 60):
            nowTime3 = i
            randomNum3 = random.random() - 0.5
            x3.append(nowTime3)
            y3.append(randomNum3)
            sheet.append([str(nowTime3)[0:8], str(randomNum3)[0:8]])
        wb7.save('倾角7.xlsx')
        cx = self.figure.add_subplot(223)
        cx.plot(x3[0:60], y3[0:60])
        cx.set_ylabel('角度')
        cx.set_xlabel('时间/s')

        # 创建workbook
        wb8 = openpyxl.Workbook()
        sheet = wb8.active
        sheet.append(['时间', '数据'])
        # 生成数据
        x4 = []
        y4 = []
        for i in range(0, 60):
            nowTime4 = i
            randomNum4 = random.random() - 0.5
            x4.append(nowTime4)
            y4.append(randomNum4)
            sheet.append([str(nowTime4)[0:8], str(randomNum4)[0:8]])
        wb8.save('倾角8.xlsx')
        dx = self.figure.add_subplot(224)
        dx.plot(x4[0:60], y4[0:60])
        dx.set_ylabel('角度')
        dx.set_xlabel('时间/s')


class angle_2(QWidget):
    def __init__(self, parent=None):
        super(angle_2, self).__init__(parent)
        self.initUi()

    def initUi(self):
        self.layout = QVBoxLayout(self)
        self.mpl = MyMplCanvas(self, width=5, height=4, dpi=100)
        self.mpl.angle_2_plot()
        self.layout.addWidget(self.mpl)